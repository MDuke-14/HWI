from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
from typing import List, Dict

def generate_monthly_report(user_data: dict, entries: List[Dict], vacation_data: dict, month: int, year: int) -> Workbook:
    """
    Gera relatório mensal em Excel no formato da Folha de Ponto HWI
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ponto {month}-{year}"
    
    # Configurações de estilo
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws.merge_cells('A1:T1')
    ws['A1'] = f"Folha de Ponto Técnico {user_data.get('full_name', user_data['username'])}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalhos superiores (linha 2)
    ws['A2'] = 'Horas extra'
    ws['B2'] = 'Horas a descontar'
    ws['C2'] = 'Dias Esp.'
    ws['D2'] = 'Alimentação'
    ws['Q2'] = 'Ferias'
    
    for cell in ['A2', 'B2', 'C2', 'D2', 'Q2']:
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Cabeçalhos da tabela (linha 3)
    headers = ['Data', 'Dia/Semana', 'Ent', 'Sai', 'Ent', 'Sai', 'Ent', 'Sai', 'Ent', 'Sai', 
               'Horas trab.', 'SA', 'ADT', 'AC', 'Tipo Pagamento']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx)
        cell.value = header
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
    
    # Organizar entries por data
    entries_by_date = {}
    for entry in entries:
        date_str = entry['date']
        if date_str not in entries_by_date:
            entries_by_date[date_str] = []
        entries_by_date[date_str].append(entry)
    
    # Preencher dados dos dias
    row = 4
    total_regular_hours = 0
    total_overtime_hours = 0
    overtime_weekend_hours = 0
    
    # Gerar todos os dias do mês
    first_day = datetime(year, month, 1)
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    
    current_date = first_day
    while current_date <= last_day:
        date_str = current_date.strftime('%Y-%m-%d')
        day_name = current_date.strftime('%A').title()
        day_names_pt = {
            'Monday': 'Segunda-feira',
            'Tuesday': 'Terça-feira',
            'Wednesday': 'Quarta-feira',
            'Thursday': 'Quinta-feira',
            'Friday': 'Sexta-feira',
            'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        day_name_pt = day_names_pt.get(day_name, day_name)
        
        # Data
        ws.cell(row=row, column=1).value = current_date.day
        ws.cell(row=row, column=1).border = border_thin
        
        # Dia da semana
        ws.cell(row=row, column=2).value = day_name_pt
        ws.cell(row=row, column=2).border = border_thin
        
        # Se há registos para este dia
        if date_str in entries_by_date:
            day_entries = entries_by_date[date_str]
            col = 3  # Coluna Ent/Sai
            
            for entry in day_entries[:4]:  # Máximo 4 entradas/saídas por dia
                if entry.get('start_time'):
                    start_dt = datetime.fromisoformat(entry['start_time'])
                    ws.cell(row=row, column=col).value = start_dt.strftime('%H:%M')
                    ws.cell(row=row, column=col).border = border_thin
                col += 1
                
                if entry.get('end_time'):
                    end_dt = datetime.fromisoformat(entry['end_time'])
                    ws.cell(row=row, column=col).value = end_dt.strftime('%H:%M')
                    ws.cell(row=row, column=col).border = border_thin
                col += 1
            
            # Total de horas do dia
            day_total = sum(e.get('total_hours', 0) for e in day_entries)
            day_regular = sum(e.get('regular_hours', 0) for e in day_entries)
            day_overtime = sum(e.get('overtime_hours', 0) for e in day_entries)
            
            ws.cell(row=row, column=11).value = round(day_total, 2)
            ws.cell(row=row, column=11).border = border_thin
            
            total_regular_hours += day_regular
            total_overtime_hours += day_overtime
            
            # Horas extras de fim de semana
            if current_date.weekday() in [5, 6] and day_overtime > 0:
                overtime_weekend_hours += day_overtime
            
            # Tipo de Pagamento (coluna 15)
            # Verifica se alguma entrada do dia está fora de zona
            is_outside_zone = any(e.get('outside_residence_zone', False) for e in day_entries)
            if is_outside_zone:
                # Pega a descrição da localização da primeira entrada fora de zona
                location = next((e.get('location_description', '') for e in day_entries if e.get('outside_residence_zone')), '')
                payment_info = f"Ajuda de Custas - {location}"
            else:
                payment_info = "Subsídio de Alimentação"
            
            ws.cell(row=row, column=15).value = payment_info
            ws.cell(row=row, column=15).border = border_thin
        
        # Colunas SA, ADT, AC (vazias por enquanto)
        for col in [12, 13, 14]:
            ws.cell(row=row, column=col).border = border_thin
        
        row += 1
        current_date += timedelta(days=1)
    
    # Secção de resumo (após os dias)
    row += 2
    
    # Horas Trabalhadas
    ws.cell(row=row, column=1).value = "Horas Trabalhadas"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2).value = round(total_regular_hours, 2)
    row += 1
    
    # Horas extra
    ws.cell(row=row, column=1).value = "Horas extra"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2).value = round(total_overtime_hours, 2)
    row += 1
    
    # NOTAS
    ws.cell(row=row, column=1).value = "NOTAS:"
    ws.cell(row=row, column=1).font = header_font
    row += 2
    
    # Abonos/Descontos/Vencimento
    ws.cell(row=row, column=1).value = "Abonos"
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    # Ajudas de Custo (exemplo: 50€ x dias trabalhados)
    days_worked = len([e for e in entries if e.get('status') == 'completed'])
    ws.cell(row=row, column=1).value = "Ajudas de Custo"
    ws.cell(row=row, column=2).value = f"50€ x {days_worked} dias"
    ws.cell(row=row, column=3).value = 50 * days_worked
    row += 1
    
    # Subsídio Alimentação
    ws.cell(row=row, column=1).value = "Subs. Alim. Cartão"
    ws.cell(row=row, column=2).value = f"10€ x {days_worked} dias"
    ws.cell(row=row, column=3).value = 10 * days_worked
    row += 1
    
    # Trabalho Suplementar
    if total_overtime_hours > 0:
        ws.cell(row=row, column=1).value = "Trabalho Suplementar (horas seguintes)"
        ws.cell(row=row, column=2).value = f"6,90€ x {round(total_overtime_hours, 2)}h"
        ws.cell(row=row, column=3).value = round(6.90 * total_overtime_hours, 2)
        row += 1
    
    # Horas fins de semana
    if overtime_weekend_hours > 0:
        ws.cell(row=row, column=1).value = "Trabalho Suplementar dias descanso"
        ws.cell(row=row, column=2).value = f"7,53€ x {round(overtime_weekend_hours, 2)}h"
        ws.cell(row=row, column=3).value = round(7.53 * overtime_weekend_hours, 2)
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = "Descontos"
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    ws.cell(row=row, column=1).value = "Horas a descontar"
    ws.cell(row=row, column=2).value = "0"
    row += 2
    
    ws.cell(row=row, column=1).value = "Vencimeno"
    ws.cell(row=row, column=1).font = header_font
    row += 2
    
    # Informações bancárias
    ws.cell(row=row, column=1).value = "Cartão alimentação"
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    ws.cell(row=row, column=1).value = "NIB:"
    ws.cell(row=row, column=2).value = user_data.get('nib', 'N/A')
    row += 1
    
    ws.cell(row=row, column=1).value = "Nº:"
    ws.cell(row=row, column=2).value = user_data.get('card_number', 'N/A')
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 8
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['O'].width = 35  # Tipo de Pagamento
    
    return wb
