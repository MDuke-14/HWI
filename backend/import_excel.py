"""
Import time entries from old Excel format
"""
import pandas as pd
from datetime import datetime, timedelta, time
import logging

def excel_date_to_python(excel_date):
    """Convert Excel serial date to Python date"""
    # Excel's epoch is 1899-12-30 (not 1900-01-01 due to a bug they kept for compatibility)
    if isinstance(excel_date, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=excel_date)
    return excel_date

def excel_time_to_python(excel_time):
    """Convert Excel decimal time to Python time (HH:MM)"""
    if pd.isna(excel_time) or excel_time == '' or excel_time == 0:
        return None
    
    # Excel time is stored as fraction of a day
    if isinstance(excel_time, (int, float)):
        total_seconds = excel_time * 24 * 3600
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours:02d}:{minutes:02d}"
    
    return None

def parse_excel_timesheet(file_path, username="Miguel Moreira"):
    """
    Parse the old Excel timesheet format and extract time entries
    
    Returns: List of dict with structure:
    {
        'date': 'YYYY-MM-DD',
        'time_entries': [{'start_time': 'HH:MM', 'end_time': 'HH:MM'}, ...],
        'outside_residence_zone': bool,
        'location_description': str or None
    }
    """
    try:
        # Read Excel file - use openpyxl directly to get raw values
        import openpyxl
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        logging.info(f"Excel file loaded with openpyxl. Max row: {sheet.max_row}, Max col: {sheet.max_column}")
        
        entries_list = []
        
        # Iterate through all rows
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            try:
                # Look for date in first column
                date_found = None
                val = row[0] if len(row) > 0 else None
                
                # Check if it's a datetime object
                if isinstance(val, datetime):
                    date_found = val
                # Check if it's an Excel serial number
                elif isinstance(val, (int, float)) and not pd.isna(val):
                    if 43000 <= val <= 48000:
                        date_found = excel_date_to_python(val)
                
                if not date_found:
                    continue
                
                # Convert to date string
                date_str = date_found.strftime('%Y-%m-%d')
                
                # Log the row we're processing
                logging.info(f"Processing row {row_idx} for date {date_str}")
                
                # Check for location info
                location = None
                outside_zone = False
                row_str = ' '.join([str(x).upper() for x in row if x is not None and x != ''])
                
                if 'MADRID' in row_str:
                    location = 'Madrid'
                    outside_zone = True
                elif 'VALENCIA' in row_str or 'VALÊNCIA' in row_str:
                    location = 'Valencia'
                    outside_zone = True
                
                # Skip vacation days
                if 'FERIAS' in row_str or 'FÉRIAS' in row_str or 'FOLGA' in row_str:
                    logging.info(f"  Skipping vacation/rest day: {date_str}")
                    continue
                
                # Find time entries - look for decimal values between 0 and 1
                time_values = []
                raw_values_log = []
                for col_idx in range(2, len(row)):  # Start from column 2
                    val = row[col_idx]
                    
                    # Skip empty/None values
                    if val is None or val == '':
                        continue
                    
                    raw_values_log.append(f"Col{col_idx}={val} (type={type(val).__name__})")
                    
                    # Check if it's a number between 0 and 1 (time as fraction of day)
                    if isinstance(val, (int, float)):
                        if 0 < val < 1:
                            time_str = excel_time_to_python(val)
                            if time_str:
                                time_values.append(time_str)
                    # Also check for datetime.time objects (in case openpyxl converts)
                    elif isinstance(val, time):
                        time_str = val.strftime('%H:%M')
                        time_values.append(time_str)
                
                if raw_values_log and row_idx <= 10:  # Only log first 10 rows
                    logging.info(f"  Raw values: {', '.join(raw_values_log[:5])}")
                
                logging.info(f"  Found {len(time_values)} time values: {time_values}")
                
                # Create pairs of start/end times
                time_pairs = []
                i = 0
                while i < len(time_values) - 1:
                    start_time = time_values[i]
                    end_time = time_values[i + 1]
                    
                    # Validate that end > start
                    try:
                        start_h, start_m = map(int, start_time.split(':'))
                        end_h, end_m = map(int, end_time.split(':'))
                        
                        if (end_h * 60 + end_m) > (start_h * 60 + start_m):
                            time_pairs.append({
                                'start_time': start_time,
                                'end_time': end_time
                            })
                            logging.info(f"    Pair: {start_time} - {end_time}")
                    except:
                        pass
                    
                    i += 2  # Move to next pair
                
                # Only add if we found valid time entries
                if time_pairs:
                    entries_list.append({
                        'date': date_str,
                        'time_entries': time_pairs,
                        'outside_residence_zone': outside_zone,
                        'location_description': location
                    })
                    logging.info(f"  ✓ Added {len(time_pairs)} entries for {date_str}")
                else:
                    logging.info(f"  ✗ No valid time pairs found for {date_str}")
                    
            except Exception as e:
                logging.warning(f"Error processing row {row_idx}: {str(e)}")
                continue
        
        workbook.close()
        logging.info(f"Total entries found: {len(entries_list)}")
        
        return {
            'success': True,
            'entries': entries_list,
            'total_days': len(entries_list)
        }
        
    except Exception as e:
        logging.error(f"Error parsing Excel file: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return {
            'success': False,
            'error': str(e),
            'entries': []
        }
